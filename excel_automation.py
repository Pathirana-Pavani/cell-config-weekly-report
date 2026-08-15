"""
EUtranCellFDD Excel Automation
--------------------------------
Folder layout expected (script should sit in this BASE folder):

    base_folder/
        excel_automation.py   <- this script
        input_zip/            <- put your .zip file here (it contains the source Excel)
        template/              <- put your template .xlsx here (empty, pre-made columns)
        output/                <- created automatically, filled result goes here

HOW IT WORKS
1. Finds the .zip file inside input_zip/, extracts it to a temp folder.
2. Finds the .xlsx/.xls file inside the extracted content.
3. Opens the "EUtranCellFDD" sheet.
4. For each of the 12 target column names below, searches ROW 1 and ROW 2
   of the sheet (handles the case where headers are split/merged across
   two rows) to find which column it lives in.
5. Reads data rows starting at ROW 6 (rows 3, 4, 5 are skipped) for just
   those columns.
6. Makes a copy of the template workbook, matches each of the 12 target
   column names against the template's OWN headers (row 1 or 2), and
   writes data only into those matched columns, starting right after the
   template's header rows. Any extra template columns (formulas, KPI,
   etc.) are left completely untouched.
7. Saves the result into output/ with a name based on the source file.

REQUIREMENTS
    pip install openpyxl

RUN
    python excel_automation.py
"""

import os
import re
import shutil
import sys
import time
import zipfile
import tempfile
from datetime import datetime

from openpyxl import load_workbook


def log(msg):
    print(msg, flush=True)


class AutomationError(Exception):
    """Raised for expected, user-fixable problems (missing file/sheet/column).

    Used instead of sys.exit() in run_pipeline() so this module can be
    imported and driven by something long-lived (e.g. a Streamlit web app)
    without sys.exit() killing the whole host process.
    """

# --------------------------------------------------------------------------
# CONFIG - adjust here if your setup differs
# --------------------------------------------------------------------------

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_ZIP_DIR = os.path.join(BASE_DIR, "input_zip")
TEMPLATE_DIR = os.path.join(BASE_DIR, "template")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
OTHER_EXPORTS_DIR = os.path.join(BASE_DIR, "other_exports")

SHEET_NAME = "EUtranCellFDD"

# Source data: header may be on row 1 or row 2; real data starts row 6.
SOURCE_HEADER_ROWS = (1, 2)
SOURCE_DATA_START_ROW = 6

# Template: the template has its own extra columns interspersed (e.g.
# NEID_CellID, Site Name, RS power, ...) that must be left untouched, so we
# match each target column by header NAME in the template too (row 1 or 2),
# instead of assuming a fixed column order. Data is written starting on the
# row right after the template's header rows.
TEMPLATE_HEADER_ROWS = (1, 2)

# The 12 columns we need to find in the source & copy, in this exact order.
# This order also determines which column (A, B, C, ...) each one is
# written to in the output sheet.
# Extra template columns filled with formulas (not copied from source data).
# Each formula is written per-row, referencing other columns by their
# ACTUAL column letter in the template (looked up by header name), so it
# keeps working even if the template's column order changes.
FORMULA_COLUMNS = [
    {
        "header": "NEID_CellID",
        "refs": ["NE ID", "E-UTRAN FDD Cell ID"],
        "formula": lambda r, c: f"=_xlfn.CONCAT({c['NE ID']}{r},{c['E-UTRAN FDD Cell ID']}{r})",
    },
    {
        "header": "Site Name",
        "refs": ["User Label"],
        "formula": lambda r, c: f'=LEFT({c["User Label"]}{r},FIND("_L",{c["User Label"]}{r})-1)',
    },
    {
        "header": "Absolute sector",
        "refs": ["User Label"],
        "formula": lambda r, c: (
            f'=_xlfn.TEXTBEFORE({c["User Label"]}{r},"L")&""&'
            f'_xlfn.TEXTBEFORE(_xlfn.TEXTAFTER({c["User Label"]}{r},"_L"),"-")'
        ),
    },
    {
        "header": "Absolute Sector",
        "refs": ["Absolute sector"],
        "formula": lambda r, c: f"=RIGHT({c['Absolute sector']}{r},1)",
    },
    {
        "header": "Name TAG",
        "refs": ["User Label"],
        "formula": lambda r, c: f"=RIGHT({c['User Label']}{r},3)",
    },
]

# RS power (template column M) is not a straight copy from EUtranCellFDD --
# it's looked up from the source workbook's ECellEquipmentFunction sheet:
# find the row there whose "Managed Element Identification" (MOI) value
# matches this row's "Baseband Resources Configuration" value, then take
# that row's cpSpeRefSigPwr.
RS_POWER_SHEET_NAME = "ECellEquipmentFunction"
RS_POWER_JOIN_COLUMN = "Baseband Resources Configuration"  # key on our side
RS_POWER_LOOKUP_KEY_COLUMN = "Managed Element Identification"  # key in ECellEquipmentFunction
RS_POWER_LOOKUP_VALUE_COLUMN = "cpSpeRefSigPwr"  # value to pull
RS_POWER_TEMPLATE_HEADER = "RS power"  # template column M

# P_A_DTCH (dB) (template column N) is looked up from the source workbook's
# PowerControlDL sheet: build a join key there by concatenating NE ID and
# E-UTRAN FDD Cell ID (same idea as =CONCAT($D6,$F6) in that sheet), match
# it against our own NEID_CellID (=CONCAT(NE ID, E-UTRAN FDD Cell ID)), and
# pull that row's paForDTCH.
PA_DTCH_SHEET_NAME = "PowerControlDL"
PA_DTCH_LOOKUP_KEY_COLUMNS = ["NE ID", "E-UTRAN FDD Cell ID"]  # PowerControlDL columns D, F
PA_DTCH_LOOKUP_VALUE_COLUMN = "paForDTCH"
PA_DTCH_TEMPLATE_HEADER = "P_A_DTCH (dB)"  # template column N

# CellMeasGroup (template column Q) is looked up from the source workbook's
# EUtranCellMeasurement sheet: build a join key there by concatenating
# NE ID (MEID) and E-UTRAN FDD Cell ID (same idea as =CONCAT($D6,$F6) in
# that sheet), match it against our own NEID_CellID, and pull that row's
# refCellMeasGroup -- keeping only the last comma-separated segment
# (e.g. "SubNetwork=50,MEID=1000,ENBFunctionFDD=1000,CellMeasGroup=10"
# -> "CellMeasGroup=10").
CELL_MEAS_GROUP_SHEET_NAME = "EUtranCellMeasurement"
CELL_MEAS_GROUP_LOOKUP_KEY_COLUMNS = ["NE ID", "E-UTRAN FDD Cell ID"]  # sheet columns D, F
CELL_MEAS_GROUP_LOOKUP_VALUE_COLUMN = "refCellMeasGroup"
CELL_MEAS_GROUP_TEMPLATE_HEADER = "CellMeasGroup"  # template column Q


def last_comma_segment(value):
    """'SubNetwork=50,...,CellMeasGroup=10' -> 'CellMeasGroup=10'."""
    if value is None:
        return None
    text = str(value)
    return text.rsplit(",", 1)[-1].strip()


# The "Traffic @ <date>" column is filled from a SEPARATE workbook -- a
# "Cell DB Export_<email>_<timestamp>.xlsx" file dropped in other_exports/ --
# joining its "Cell Name" column against our "User Label" column. The date
# in the template's header text is refreshed to the export's own date
# (parsed from the filename), since it changes every time a new export is
# used.
CELL_DB_EXPORT_SHEET_NAME = "CellAntennaPortsMap"
CELL_DB_EXPORT_KEY_COLUMN = "Cell Name"
CELL_DB_EXPORT_VALUE_COLUMN = "Traffic"
CELL_DB_EXPORT_HEADER_ROWS = (1,)
CELL_DB_EXPORT_DATA_START_ROW = 2
TRAFFIC_HEADER_PREFIX = "Traffic @"


def find_cell_db_export_file(folder):
    """Find the first 'Cell DB Export*.xlsx' file in `folder`, if any."""
    if not os.path.isdir(folder):
        return None
    for fname in os.listdir(folder):
        if fname.startswith("~$"):
            continue
        if fname.lower().endswith((".xlsx", ".xlsm")) and fname.lower().startswith("cell db export"):
            return os.path.join(folder, fname)
    return None


def extract_export_date_label(filename):
    """
    Pull an embedded YYYYMMDD (optionally followed by more digits, e.g. a
    time) out of a filename and format it like the template's own
    convention, e.g. "16-July". Falls back to today's date if none found.
    """
    m = re.search(r"(20\d{2})(\d{2})(\d{2})", filename)
    if m:
        dt = datetime.strptime(m.group(0)[:8], "%Y%m%d")
    else:
        dt = datetime.now()
    return dt.strftime("%d-%B")


def find_column_by_prefix(row_grids, merged_ranges, header_rows, prefix, max_col):
    """Return the 1-based column of the first header cell starting with `prefix` (case-insensitive)."""
    prefix_norm = normalize(prefix)
    for row in header_rows:
        for col in range(1, max_col + 1):
            val = effective_header_value(row_grids, merged_ranges, row, col)
            if val is None:
                continue
            if normalize(val).startswith(prefix_norm):
                return col
    return None

# "FDD Power (W)Authorization Value" and "LTE FDD(W) Configuration Value"
# are looked up from a SEPARATE workbook -- a "powerlicense <date>.xlsx"
# file dropped in other_exports/ -- joining its "NE ID" column (one row per
# NE, eNodeB-level) against our own "NE ID" column.
POWER_LICENSE_SHEET_NAME = "Sheet1"
POWER_LICENSE_KEY_COLUMN = "NE ID"
POWER_LICENSE_AUTH_COLUMN = "Authorization Value"
POWER_LICENSE_CONFIG_COLUMN = "Configuration Value"
POWER_LICENSE_HEADER_ROWS = (1,)
POWER_LICENSE_DATA_START_ROW = 2
FDD_POWER_AUTH_TEMPLATE_HEADER = "FDD Power (W)Authorization Value"
LTE_FDD_CONFIG_TEMPLATE_HEADER = "LTE FDD(W) Configuration Value"


def find_power_license_file(folder):
    """Find the first 'powerlicense*.xlsx' file in `folder`, if any."""
    if not os.path.isdir(folder):
        return None
    for fname in os.listdir(folder):
        if fname.startswith("~$"):
            continue
        if fname.lower().endswith((".xlsx", ".xlsm")) and fname.lower().startswith("powerlicense"):
            return os.path.join(folder, fname)
    return None


TARGET_COLUMNS = [
    "SubNetwork ID",
    "NE ID",
    "E-UTRAN FDD Cell ID",
    "User Label",
    "Managed Element Identification",
    "Baseband Resources Configuration",
    "Frequency Band Indicator(Effective when Cell Reset)",
    "PCI(Effective when Cell Reset)",
    "Signal Power Ratio (PB)",
    "Number of Tx Antenna Port in Cell(Effective when Cell Reset)",
    "DL System Bandwidth (MHz)(Effective when Cell Reset)",
    "Maximum RBs Allocated for Downlink UE",
]


# --------------------------------------------------------------------------
# HELPERS
# --------------------------------------------------------------------------

def normalize(text):
    """Lowercase, strip, and collapse whitespace for fuzzy header matching."""
    if text is None:
        return ""
    text = str(text)
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text.lower()


MAX_SCAN_COLS = 300  # these exports often report a bogus/tiny sheet dimension,
                      # so we must always pass an explicit max_col/max_row to
                      # iter_rows() or the read-only iterator yields nothing.


def fetch_header_rows(ws, header_rows, max_col=MAX_SCAN_COLS):
    """
    Read only the given header rows into memory as {row_num: tuple_of_values}.
    Works with read_only worksheets (fast, no full-sheet load needed).
    """
    min_r, max_r = min(header_rows), max(header_rows)
    grids = {}
    for i, row_values in enumerate(
        ws.iter_rows(min_row=min_r, max_row=max_r, max_col=max_col, values_only=True),
        start=min_r,
    ):
        grids[i] = row_values
    return grids


def get_merged_ranges(ws):
    """Safely fetch merged cell ranges (works for read_only sheets too, in most openpyxl versions)."""
    try:
        return list(ws.merged_cells.ranges)
    except Exception:
        return []


def effective_header_value(row_grids, merged_ranges, row, col):
    """Resolve a header cell value, using merged-range top-left value if needed."""
    values = row_grids.get(row)
    if values and col - 1 < len(values) and values[col - 1] is not None:
        return values[col - 1]
    for merged_range in merged_ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
            top_row, top_col = merged_range.min_row, merged_range.min_col
            top_values = row_grids.get(top_row)
            if top_values and top_col - 1 < len(top_values):
                return top_values[top_col - 1]
    return None


def _exact_key(text):
    """Whitespace-trimmed but case-preserving key, for exact header matching."""
    if text is None:
        return ""
    text = str(text).replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip()


def find_header_columns_fast(row_grids, merged_ranges, header_rows, target_columns, max_col):
    """
    Search the given header rows (pre-fetched as tuples) for each target
    column name. Returns dict: target_name -> column_index (1-based).
    Missing -> None.

    Two passes: exact (case-sensitive) match first, since some sheets have
    distinct headers that differ only by case (e.g. "Absolute sector" vs
    "Absolute Sector") -- a case-insensitive-only match would collide the
    two. Any target still unmatched after pass 1 falls back to a
    case-insensitive match in pass 2.
    """
    found = {name: None for name in target_columns}

    # Pass 1: exact, case-sensitive.
    exact_targets = {_exact_key(name): name for name in target_columns}
    for row in header_rows:
        for col in range(1, max_col + 1):
            val = effective_header_value(row_grids, merged_ranges, row, col)
            if val is None:
                continue
            key = _exact_key(val)
            if key in exact_targets:
                original_name = exact_targets[key]
                if found[original_name] is None:
                    found[original_name] = col

    # Pass 2: case-insensitive fallback, only for names still unmatched.
    remaining = [name for name, col in found.items() if col is None]
    if remaining:
        normalized_targets = {normalize(name): name for name in remaining}
        for row in header_rows:
            for col in range(1, max_col + 1):
                val = effective_header_value(row_grids, merged_ranges, row, col)
                if val is None:
                    continue
                norm_val = normalize(val)
                if norm_val in normalized_targets:
                    original_name = normalized_targets[norm_val]
                    if found[original_name] is None:
                        found[original_name] = col

    return found


def build_lookup_dict(wb, sheet_name, key_header, value_header,
                       header_rows=SOURCE_HEADER_ROWS, data_start_row=SOURCE_DATA_START_ROW):
    """
    Read `sheet_name` from an already-open workbook and return a dict
    mapping key_header's value -> value_header's value, one entry per data
    row (last one wins if a key repeats).
    """
    sheet_actual_name = find_sheet(wb, sheet_name)
    if sheet_actual_name is None:
        log(f"WARNING: Sheet '{sheet_name}' not found; lookups against it will be blank.")
        return {}
    ws = wb[sheet_actual_name]

    header_grids = fetch_header_rows(ws, header_rows)
    merged_ranges = get_merged_ranges(ws)
    col_map = find_header_columns_fast(
        header_grids, merged_ranges, header_rows, [key_header, value_header], MAX_SCAN_COLS
    )
    key_col, value_col = col_map[key_header], col_map[value_header]
    if key_col is None or value_col is None:
        log(f"WARNING: Could not find '{key_header}' and/or '{value_header}' in "
            f"sheet '{sheet_name}'; lookups against it will be blank.")
        return {}
    key_idx, value_idx = key_col - 1, value_col - 1
    max_needed_col = max(key_idx, value_idx) + 1

    lookup = {}
    for row_tuple in ws.iter_rows(
        min_row=data_start_row,
        max_row=data_start_row + 500_000,
        max_col=max_needed_col,
        values_only=True,
    ):
        if all(v is None for v in row_tuple):
            break
        key_val = row_tuple[key_idx] if key_idx < len(row_tuple) else None
        if key_val is None:
            continue
        lookup[key_val] = row_tuple[value_idx] if value_idx < len(row_tuple) else None

    return lookup


def build_multikey_lookup_dict(wb, sheet_name, key_headers, value_header,
                                header_rows=SOURCE_HEADER_ROWS, data_start_row=SOURCE_DATA_START_ROW,
                                value_transform=None):
    """
    Like build_lookup_dict, but the key is formed by concatenating several
    columns (as strings, in the given order) -- e.g. NE ID + E-UTRAN FDD
    Cell ID, matching a sheet's own "=CONCAT(colA, colB)" convention.
    Returns dict: concatenated_key_string -> value_header's value.
    """
    sheet_actual_name = find_sheet(wb, sheet_name)
    if sheet_actual_name is None:
        log(f"WARNING: Sheet '{sheet_name}' not found; lookups against it will be blank.")
        return {}
    ws = wb[sheet_actual_name]

    header_grids = fetch_header_rows(ws, header_rows)
    merged_ranges = get_merged_ranges(ws)
    needed = key_headers + [value_header]
    col_map = find_header_columns_fast(header_grids, merged_ranges, header_rows, needed, MAX_SCAN_COLS)

    missing = [h for h in needed if col_map.get(h) is None]
    if missing:
        log(f"WARNING: Could not find {missing} in sheet '{sheet_name}'; "
            f"lookups against it will be blank.")
        return {}

    key_idxs = [col_map[h] - 1 for h in key_headers]
    value_idx = col_map[value_header] - 1
    max_needed_col = max(key_idxs + [value_idx]) + 1

    lookup = {}
    for row_tuple in ws.iter_rows(
        min_row=data_start_row,
        max_row=data_start_row + 500_000,
        max_col=max_needed_col,
        values_only=True,
    ):
        if all(v is None for v in row_tuple):
            break
        key_parts = [row_tuple[idx] if idx < len(row_tuple) else None for idx in key_idxs]
        if any(part is None for part in key_parts):
            continue
        key = "".join(str(part) for part in key_parts)
        raw_value = row_tuple[value_idx] if value_idx < len(row_tuple) else None
        lookup[key] = value_transform(raw_value) if value_transform else raw_value

    return lookup


def find_zip_file(folder):
    for fname in os.listdir(folder):
        if fname.lower().endswith(".zip"):
            return os.path.join(folder, fname)
    return None


def find_excel_file(folder):
    """Recursively search a folder for the first .xlsx/.xlsm/.xls file."""
    for root, _dirs, files in os.walk(folder):
        for fname in files:
            if fname.lower().endswith((".xlsx", ".xlsm", ".xls")) and not fname.startswith("~$"):
                return os.path.join(root, fname)
    return None


def find_template_file(folder):
    for fname in os.listdir(folder):
        if fname.lower().endswith((".xlsx", ".xlsm")) and not fname.startswith("~$"):
            return os.path.join(folder, fname)
    return None


def find_sheet(wb, target_name):
    """Case-insensitive sheet name lookup."""
    for name in wb.sheetnames:
        if name.strip().lower() == target_name.strip().lower():
            return name
    # fallback: partial match
    for name in wb.sheetnames:
        if target_name.strip().lower() in name.strip().lower():
            return name
    return None


# --------------------------------------------------------------------------
# PIPELINE (importable -- used by both the CLI below and webapp.py)
# --------------------------------------------------------------------------

def run_pipeline(zip_path, template_path, other_exports_dir, output_dir,
                  cell_db_export_path=None, power_license_path=None):
    """
    Run the full extraction/lookup/fill pipeline and return
    (output_path, summary_dict). Raises AutomationError on any expected,
    user-fixable problem (missing file/sheet/column) instead of exiting the
    process, so this is safe to call from a long-lived host like a web app.

    cell_db_export_path / power_license_path: pass these explicitly to use
    a specific file regardless of its name (e.g. a web upload with an
    arbitrary filename). If omitted, they're auto-discovered by name inside
    other_exports_dir, same as the CLI's folder-drop workflow.
    """
    os.makedirs(output_dir, exist_ok=True)
    summary = {"warnings": []}

    def note(msg):
        summary["warnings"].append(msg)
        log(msg)

    # 1. Extract the zip -------------------------------------------------
    temp_dir = tempfile.mkdtemp(prefix="eutrancell_")
    with zipfile.ZipFile(zip_path, "r") as zf:
        zf.extractall(temp_dir)
    log(f"Extracted to: {temp_dir}")

    # 2. Locate the source Excel file ----------------------------------------------
    source_excel_path = find_excel_file(temp_dir)
    if not source_excel_path:
        raise AutomationError("No Excel file found inside the extracted zip contents.")
    source_filename = os.path.basename(source_excel_path)
    size_mb = os.path.getsize(source_excel_path) / (1024 * 1024)
    log(f"Found source Excel: {source_filename} ({size_mb:.1f} MB)")
    summary["source_filename"] = source_filename

    # 3. Confirm the template exists ---------------------------------------------
    if not template_path or not os.path.isfile(template_path):
        raise AutomationError(f"Template file not found: '{template_path}'.")
    log(f"Using template: {template_path}")

    # 4. Load source workbook & sheet -------------------------------------------
    # read_only=True avoids openpyxl building an in-memory model of every sheet
    # in the workbook up front, which is what causes big radio-config exports
    # (often many sheets, heavy formatting) to appear "stuck" for a long time.
    log("Loading source workbook (read-only/streaming mode)... this can still "
        "take a bit for large files, please wait.")
    t0 = time.time()
    src_wb = load_workbook(source_excel_path, data_only=True, read_only=True)
    log(f"Workbook opened in {time.time() - t0:.1f}s.")

    src_sheet_name = find_sheet(src_wb, SHEET_NAME)
    if not src_sheet_name:
        raise AutomationError(
            f"Sheet '{SHEET_NAME}' not found in source file. "
            f"Available sheets: {src_wb.sheetnames}"
        )
    src_ws = src_wb[src_sheet_name]
    log(f"Using source sheet: '{src_sheet_name}'")

    # 5. Find target columns in source headers -----------------------------------
    log("Reading source header rows and matching the 12 target columns...")
    header_grids = fetch_header_rows(src_ws, SOURCE_HEADER_ROWS)
    merged_ranges = get_merged_ranges(src_ws)
    src_col_map = find_header_columns_fast(
        header_grids, merged_ranges, SOURCE_HEADER_ROWS, TARGET_COLUMNS, MAX_SCAN_COLS
    )

    missing_in_source = [name for name, col in src_col_map.items() if col is None]
    if missing_in_source:
        note("These columns were not found in the source header rows "
             f"{SOURCE_HEADER_ROWS} and will be left blank in the output: "
             + ", ".join(missing_in_source))
    else:
        log(f"All {len(TARGET_COLUMNS)} target columns matched in source headers.")
    summary["missing_in_source"] = missing_in_source

    # 6. Read data rows from source (single streaming pass) -----------------------
    log(f"Reading data rows starting at row {SOURCE_DATA_START_ROW} (streaming)...")
    t0 = time.time()
    # Map: target column name -> 0-based tuple index, for fast lookup per row
    name_to_index = {name: (col - 1) for name, col in src_col_map.items() if col is not None}

    max_needed_col = max((col for col in name_to_index.values()), default=-1) + 1

    rows_data = []
    for row_tuple in src_ws.iter_rows(
        min_row=SOURCE_DATA_START_ROW,
        max_row=SOURCE_DATA_START_ROW + 500_000,
        max_col=max_needed_col,
        values_only=True,
    ):
        if all(v is None for v in row_tuple):  # end of data
            break
        row_values = {}
        any_value = False
        for name in TARGET_COLUMNS:
            idx = name_to_index.get(name)
            if idx is None or idx >= len(row_tuple):
                row_values[name] = None
                continue
            val = row_tuple[idx]
            row_values[name] = val
            if val is not None and str(val).strip() != "":
                any_value = True
        if any_value:
            rows_data.append(row_values)

    log(f"Read {len(rows_data)} data rows from source in {time.time() - t0:.1f}s.")
    summary["rows_read"] = len(rows_data)

    # 6b. Build the RS power lookup from ECellEquipmentFunction and apply it --------
    log(f"Building RS power lookup from '{RS_POWER_SHEET_NAME}' sheet "
        f"({RS_POWER_LOOKUP_KEY_COLUMN} -> {RS_POWER_LOOKUP_VALUE_COLUMN})...")
    rs_power_lookup = build_lookup_dict(
        src_wb, RS_POWER_SHEET_NAME, RS_POWER_LOOKUP_KEY_COLUMN, RS_POWER_LOOKUP_VALUE_COLUMN
    )
    log(f"RS power lookup has {len(rs_power_lookup)} entries.")

    rs_power_misses = 0
    for row_values in rows_data:
        join_val = row_values.get(RS_POWER_JOIN_COLUMN)
        rs_power = rs_power_lookup.get(join_val)
        row_values[RS_POWER_TEMPLATE_HEADER] = rs_power
        if rs_power is None:
            rs_power_misses += 1
    if rs_power_misses:
        note(f"{rs_power_misses} row(s) had no matching '{RS_POWER_JOIN_COLUMN}' "
             f"value in '{RS_POWER_SHEET_NAME}' -- '{RS_POWER_TEMPLATE_HEADER}' left blank for those.")
    summary["rs_power"] = {"entries": len(rs_power_lookup), "misses": rs_power_misses}

    # 6c. Build the P_A_DTCH lookup from PowerControlDL and apply it ---------------
    log(f"Building P_A_DTCH lookup from '{PA_DTCH_SHEET_NAME}' sheet "
        f"({'+'.join(PA_DTCH_LOOKUP_KEY_COLUMNS)} -> {PA_DTCH_LOOKUP_VALUE_COLUMN})...")
    pa_dtch_lookup = build_multikey_lookup_dict(
        src_wb, PA_DTCH_SHEET_NAME, PA_DTCH_LOOKUP_KEY_COLUMNS, PA_DTCH_LOOKUP_VALUE_COLUMN
    )
    log(f"P_A_DTCH lookup has {len(pa_dtch_lookup)} entries.")

    pa_dtch_misses = 0
    for row_values in rows_data:
        key_parts = [row_values.get(h) for h in PA_DTCH_LOOKUP_KEY_COLUMNS]
        if any(part is None for part in key_parts):
            pa_dtch_value = None
        else:
            join_key = "".join(str(part) for part in key_parts)
            pa_dtch_value = pa_dtch_lookup.get(join_key)
        row_values[PA_DTCH_TEMPLATE_HEADER] = pa_dtch_value
        if pa_dtch_value is None:
            pa_dtch_misses += 1
    if pa_dtch_misses:
        note(f"{pa_dtch_misses} row(s) had no matching "
             f"{'+'.join(PA_DTCH_LOOKUP_KEY_COLUMNS)} key in '{PA_DTCH_SHEET_NAME}' -- "
             f"'{PA_DTCH_TEMPLATE_HEADER}' left blank for those.")
    summary["pa_dtch"] = {"entries": len(pa_dtch_lookup), "misses": pa_dtch_misses}

    # 6d. Build the CellMeasGroup lookup from EUtranCellMeasurement and apply it ----
    log(f"Building CellMeasGroup lookup from '{CELL_MEAS_GROUP_SHEET_NAME}' sheet "
        f"({'+'.join(CELL_MEAS_GROUP_LOOKUP_KEY_COLUMNS)} -> {CELL_MEAS_GROUP_LOOKUP_VALUE_COLUMN})...")
    cell_meas_group_lookup = build_multikey_lookup_dict(
        src_wb, CELL_MEAS_GROUP_SHEET_NAME, CELL_MEAS_GROUP_LOOKUP_KEY_COLUMNS,
        CELL_MEAS_GROUP_LOOKUP_VALUE_COLUMN, value_transform=last_comma_segment,
    )
    log(f"CellMeasGroup lookup has {len(cell_meas_group_lookup)} entries.")

    cell_meas_group_misses = 0
    for row_values in rows_data:
        key_parts = [row_values.get(h) for h in CELL_MEAS_GROUP_LOOKUP_KEY_COLUMNS]
        if any(part is None for part in key_parts):
            cell_meas_group_value = None
        else:
            join_key = "".join(str(part) for part in key_parts)
            cell_meas_group_value = cell_meas_group_lookup.get(join_key)
        row_values[CELL_MEAS_GROUP_TEMPLATE_HEADER] = cell_meas_group_value
        if cell_meas_group_value is None:
            cell_meas_group_misses += 1
    if cell_meas_group_misses:
        note(f"{cell_meas_group_misses} row(s) had no matching "
             f"{'+'.join(CELL_MEAS_GROUP_LOOKUP_KEY_COLUMNS)} key in "
             f"'{CELL_MEAS_GROUP_SHEET_NAME}' -- "
             f"'{CELL_MEAS_GROUP_TEMPLATE_HEADER}' left blank for those.")
    summary["cell_meas_group"] = {"entries": len(cell_meas_group_lookup), "misses": cell_meas_group_misses}

    src_wb.close()

    # 6e. Build the Traffic lookup from the separate Cell DB Export workbook -------
    traffic_label = None
    traffic_lookup = {}
    summary["traffic"] = None
    cell_db_path = cell_db_export_path or find_cell_db_export_file(other_exports_dir)
    if cell_db_path is None:
        note(f"No 'Cell DB Export*.xlsx' file found in '{other_exports_dir}' -- "
             f"the '{TRAFFIC_HEADER_PREFIX}' column will be left as-is.")
    else:
        cell_db_filename = os.path.basename(cell_db_path)
        traffic_label = extract_export_date_label(cell_db_filename)
        log(f"Found Cell DB Export: {cell_db_filename} (date label: {traffic_label})")
        cell_db_wb = load_workbook(cell_db_path, data_only=True, read_only=True)
        traffic_lookup = build_lookup_dict(
            cell_db_wb, CELL_DB_EXPORT_SHEET_NAME, CELL_DB_EXPORT_KEY_COLUMN, CELL_DB_EXPORT_VALUE_COLUMN,
            header_rows=CELL_DB_EXPORT_HEADER_ROWS, data_start_row=CELL_DB_EXPORT_DATA_START_ROW,
        )
        cell_db_wb.close()
        log(f"Traffic lookup has {len(traffic_lookup)} entries.")

        traffic_misses = 0
        for row_values in rows_data:
            traffic_value = traffic_lookup.get(row_values.get("User Label"))
            row_values["__TRAFFIC__"] = traffic_value
            if traffic_value is None:
                traffic_misses += 1
        if traffic_misses:
            note(f"{traffic_misses} row(s) had no matching User Label in "
                 f"'{CELL_DB_EXPORT_SHEET_NAME}' -- Traffic left blank for those.")
        summary["traffic"] = {
            "file": cell_db_filename, "label": traffic_label,
            "entries": len(traffic_lookup), "misses": traffic_misses,
        }

    # 6f. Build the FDD power lookups from the separate power license workbook -----
    summary["power_license"] = None
    power_license_path = power_license_path or find_power_license_file(other_exports_dir)
    if power_license_path is None:
        note(f"No 'powerlicense*.xlsx' file found in '{other_exports_dir}' -- "
             f"'{FDD_POWER_AUTH_TEMPLATE_HEADER}' and '{LTE_FDD_CONFIG_TEMPLATE_HEADER}' "
             f"will be left blank.")
    else:
        power_license_filename = os.path.basename(power_license_path)
        log(f"Found power license export: {power_license_filename}")
        power_wb = load_workbook(power_license_path, data_only=True, read_only=True)
        auth_lookup = build_lookup_dict(
            power_wb, POWER_LICENSE_SHEET_NAME, POWER_LICENSE_KEY_COLUMN, POWER_LICENSE_AUTH_COLUMN,
            header_rows=POWER_LICENSE_HEADER_ROWS, data_start_row=POWER_LICENSE_DATA_START_ROW,
        )
        config_lookup = build_lookup_dict(
            power_wb, POWER_LICENSE_SHEET_NAME, POWER_LICENSE_KEY_COLUMN, POWER_LICENSE_CONFIG_COLUMN,
            header_rows=POWER_LICENSE_HEADER_ROWS, data_start_row=POWER_LICENSE_DATA_START_ROW,
        )
        power_wb.close()
        # Keys in the license file are numeric NE IDs; our rows' NE ID may be
        # text -- normalize both sides to strings so they compare equal.
        auth_lookup = {str(k): v for k, v in auth_lookup.items()}
        config_lookup = {str(k): v for k, v in config_lookup.items()}
        log(f"Power license lookup has {len(auth_lookup)} NE ID entries.")

        power_misses = 0
        for row_values in rows_data:
            ne_id_key = row_values.get("NE ID")
            ne_id_key = str(ne_id_key) if ne_id_key is not None else None
            auth_value = auth_lookup.get(ne_id_key) if ne_id_key is not None else None
            config_value = config_lookup.get(ne_id_key) if ne_id_key is not None else None
            row_values[FDD_POWER_AUTH_TEMPLATE_HEADER] = auth_value
            row_values[LTE_FDD_CONFIG_TEMPLATE_HEADER] = config_value
            if auth_value is None and config_value is None:
                power_misses += 1
        if power_misses:
            note(f"{power_misses} row(s) had no matching NE ID in "
                 f"'{power_license_filename}' -- power values left blank for those.")
        summary["power_license"] = {
            "file": power_license_filename, "entries": len(auth_lookup), "misses": power_misses,
        }

    # 7. Prepare output workbook from template ------------------------------------
    output_name = f"{os.path.splitext(source_filename)[0]}_filled.xlsx"
    output_path = os.path.join(output_dir, output_name)
    shutil.copy(template_path, output_path)

    out_wb = load_workbook(output_path)
    out_sheet_name = find_sheet(out_wb, SHEET_NAME) or out_wb.sheetnames[0]
    out_ws = out_wb[out_sheet_name]

    # Match each target column by NAME against the template's own headers
    # (row 1 or 2) -- the template has extra columns (formulas, KPI, etc.)
    # interspersed that must be left completely untouched.
    tpl_header_grids = fetch_header_rows(out_ws, TEMPLATE_HEADER_ROWS, max_col=out_ws.max_column)
    tpl_merged_ranges = get_merged_ranges(out_ws)
    DATA_COLUMNS = TARGET_COLUMNS + [
        RS_POWER_TEMPLATE_HEADER, PA_DTCH_TEMPLATE_HEADER, CELL_MEAS_GROUP_TEMPLATE_HEADER,
        FDD_POWER_AUTH_TEMPLATE_HEADER, LTE_FDD_CONFIG_TEMPLATE_HEADER,
    ]
    tpl_col_map = find_header_columns_fast(
        tpl_header_grids, tpl_merged_ranges, TEMPLATE_HEADER_ROWS, DATA_COLUMNS, out_ws.max_column
    )

    # Also locate the formula-column headers and any extra ref columns they
    # need (e.g. "Absolute sector" is both a formula target and a ref for
    # "Absolute Sector"), so every name used below resolves to a column letter.
    formula_headers = [fc["header"] for fc in FORMULA_COLUMNS]
    all_formula_names = set(formula_headers) | {ref for fc in FORMULA_COLUMNS for ref in fc["refs"]}
    names_to_locate = sorted(all_formula_names - set(DATA_COLUMNS))
    formula_col_map = find_header_columns_fast(
        tpl_header_grids, tpl_merged_ranges, TEMPLATE_HEADER_ROWS, names_to_locate, out_ws.max_column
    )
    combined_col_map = {**tpl_col_map, **formula_col_map}

    missing_formula_cols = [name for name in all_formula_names if combined_col_map.get(name) is None]
    if missing_formula_cols:
        note("These formula-related columns were not found in the template "
             "and their formulas will be skipped: " + ", ".join(missing_formula_cols))

    from openpyxl.utils import get_column_letter

    col_letters = {
        name: get_column_letter(col)
        for name, col in combined_col_map.items()
        if col is not None
    }

    active_formula_columns = [
        fc for fc in FORMULA_COLUMNS
        if combined_col_map.get(fc["header"]) is not None
        and all(combined_col_map.get(ref) is not None for ref in fc["refs"])
    ]

    # Locate the "Traffic @ <date>" column by prefix (its exact date text
    # changes with each export) and refresh its header to this export's date.
    traffic_col = find_column_by_prefix(
        tpl_header_grids, tpl_merged_ranges, TEMPLATE_HEADER_ROWS, TRAFFIC_HEADER_PREFIX, out_ws.max_column
    )
    if traffic_col is None:
        note(f"No '{TRAFFIC_HEADER_PREFIX} ...' column found in the template -- "
             f"Traffic will not be written.")
    elif traffic_label is not None:
        header_row_to_update = max(
            (r for r in TEMPLATE_HEADER_ROWS
             if effective_header_value(tpl_header_grids, tpl_merged_ranges, r, traffic_col) is not None),
            default=min(TEMPLATE_HEADER_ROWS),
        )
        out_ws.cell(row=header_row_to_update, column=traffic_col, value=f"{TRAFFIC_HEADER_PREFIX} {traffic_label}")

    # Data should start right after the LAST header row that actually has
    # content (some templates only use row 1, leaving row 2 blank).
    template_data_start_row = min(TEMPLATE_HEADER_ROWS) + 1
    for row_num in sorted(TEMPLATE_HEADER_ROWS):
        row_vals = tpl_header_grids.get(row_num) or ()
        if any(v is not None and str(v).strip() != "" for v in row_vals):
            template_data_start_row = row_num + 1
    missing_in_template = [name for name, col in tpl_col_map.items() if col is None]
    if missing_in_template:
        note("These columns were not found in the template headers and will "
             "be skipped: " + ", ".join(missing_in_template))
    summary["missing_in_template"] = missing_in_template

    log(f"Writing into template sheet: '{out_sheet_name}' "
        f"(matched by header name, starting at row {template_data_start_row})")

    # 8. Write data into template, only into the columns matched by name
    for i, row_values in enumerate(rows_data):
        out_row = template_data_start_row + i
        for name in DATA_COLUMNS:
            out_col = tpl_col_map.get(name)
            if out_col is None:
                continue
            out_ws.cell(row=out_row, column=out_col, value=row_values.get(name))
        if traffic_col is not None:
            out_ws.cell(row=out_row, column=traffic_col, value=row_values.get("__TRAFFIC__"))
        for fc in active_formula_columns:
            target_col = combined_col_map[fc["header"]]
            out_ws.cell(row=out_row, column=target_col, value=fc["formula"](out_row, col_letters))

    out_wb.save(output_path)
    log(f"\nDone. Output saved to: {output_path}")
    summary["output_path"] = output_path
    summary["output_name"] = output_name
    summary["traffic_label"] = traffic_label

    # cleanup temp extraction
    shutil.rmtree(temp_dir, ignore_errors=True)

    return output_path, summary


# --------------------------------------------------------------------------
# CLI ENTRY POINT
# --------------------------------------------------------------------------

def main():
    zip_path = find_zip_file(INPUT_ZIP_DIR)
    if not zip_path:
        sys.exit(f"ERROR: No .zip file found in '{INPUT_ZIP_DIR}'.")
    log(f"Found zip: {zip_path}")

    template_path = find_template_file(TEMPLATE_DIR)
    if not template_path:
        sys.exit(f"ERROR: No template .xlsx file found in '{TEMPLATE_DIR}'.")

    try:
        run_pipeline(zip_path, template_path, OTHER_EXPORTS_DIR, OUTPUT_DIR)
    except AutomationError as e:
        sys.exit(f"ERROR: {e}")


if __name__ == "__main__":
    main()