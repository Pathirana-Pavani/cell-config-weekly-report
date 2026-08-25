"""
huawei_automation.py
---------------------
Huawei site report automation. Unlike the ZTE pipeline (one zip with a
multi-sheet workbook), Huawei exports are THREE SEPARATE files that the
user uploads individually:

    1. LST CELL              -- the "spine": most columns are copied
                                 straight from here.
    2. LSTPDSCH               -- source for Reference signal power(0.1dBm)
                                 and PB.
    3. LST CELLDLPCPDSCHPA    -- source for PA for even power
                                 distribution(dB).

Each export's "result" sheet has: row 1 = category, row 2 = friendly
column name, data starting row 3 (some rows are base-station-only rows
with no cell data -- e.g. Base Station Name filled but Cell Name blank --
these are skipped).

JOIN KEY: Cell ID + Base Station Name (matching the reference mapping
file's own `=CONCATENATE(D,B)` convention), built the same way from each
of the three files even though their own column order differs (LST CELL:
Base Station Name then Local Cell ID; the other two: same, but the
manually-built reference used "concatenate B,A" to land on the same
ID-then-name order).

Written into the template (template/huawei_configuration.xlsx, itself a
stripped-down copy of the manually-built "huawei configuration
mapping.xlsx" with the sample data removed but headers/colors kept):
  - Direct-copy columns: literal values from LST CELL.
  - Absolute sector + the blank "join key" column: Excel FORMULAS
    (self-contained, single-row, no external file dependency) -- same
    convention as excel_automation.py's FORMULA_COLUMNS.
  - Reference signal power / PB / PA for even power distribution: literal
    Python-computed values (the actual cross-file lookup), since relying
    on live external-workbook formulas is exactly the fragility this
    automation replaces.
"""

import os
import shutil
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from excel_automation import (
    AutomationError,
    MAX_SCAN_COLS,
    fetch_header_rows,
    find_header_columns_fast,
    find_sheet,
    get_merged_ranges,
    log,
)

SOURCE_SHEET_NAME = "result"
SOURCE_HEADER_ROWS = (1, 2)
SOURCE_DATA_START_ROW = 3

TEMPLATE_HEADER_ROWS = (1,)
TEMPLATE_DATA_START_ROW = 2

# Template header name -> LST CELL source header name (identical except Cell ID).
LST_CELL_DIRECT_COLUMNS = {
    "Base Station Name": "Base Station Name",
    "Cell ID": "Local Cell ID",
    "Cell Name": "Cell Name",
    "Cell active state": "Cell active state",
    "Frequency band": "Frequency band",
    "Physical cell ID": "Physical cell ID",
    "Downlink EARFCN": "Downlink EARFCN",
    "Downlink bandwidth": "Downlink bandwidth",
    "Uplink bandwidth": "Uplink bandwidth",
    "CRS Port Number": "CRS Port Number",
    "Cell transmission and reception mode": "Cell transmission and reception mode",
}

# Join key columns in LSTPDSCH / LST CELLDLPCPDSCHPA: Local cell ID then
# Base Station Name (ID-then-name, matching LST CELL's Cell ID + Base
# Station Name order).
JOIN_KEY_SOURCE_COLUMNS = ["Local cell ID", "Base Station Name"]

LSTPDSCH_VALUE_COLUMNS = ["Reference signal power(0.1dBm)", "PB"]
CELLDLPCPDSCHPA_VALUE_COLUMNS = ["PA for even power distribution(dB)"]

TEMPLATE_DIRECT_COLUMNS = list(LST_CELL_DIRECT_COLUMNS.keys())
TEMPLATE_LOOKUP_COLUMNS = LSTPDSCH_VALUE_COLUMNS + CELLDLPCPDSCHPA_VALUE_COLUMNS
TEMPLATE_JOIN_KEY_COLUMN = "Cell ID + Base Station Name"


def _find(needle, haystack, start=1):
    idx = haystack.find(needle, start - 1)
    return None if idx == -1 else idx + 1


def absolute_sector_formula(cell_name_ref):
    """
    1/2/3 stay as-is, 4-8 shift down to 1-5, anything else -> blank --
    same remap as ZTE's Absolute Sector, applied to the digit right after
    "_L" in Cell Name (Huawei's own convention is simpler than ZTE's:
    just the one digit via MID, no further text needed).
    """
    d = f'MID({cell_name_ref},FIND("_L",{cell_name_ref})+2,1)'
    return (
        f"=IFERROR(IF(AND(VALUE({d})>=1,VALUE({d})<=3),VALUE({d}),"
        f'IF(AND(VALUE({d})>=4,VALUE({d})<=8),VALUE({d})-3,"")),"")'
    )


def _read_source_sheet(path, needed_columns):
    """Open a Huawei export file's 'result' sheet and return (ws, col_map, wb)."""
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_name = find_sheet(wb, SOURCE_SHEET_NAME) or wb.sheetnames[0]
    ws = wb[sheet_name]
    header_grids = fetch_header_rows(ws, SOURCE_HEADER_ROWS)
    merged = get_merged_ranges(ws)
    col_map = find_header_columns_fast(header_grids, merged, SOURCE_HEADER_ROWS, needed_columns, MAX_SCAN_COLS)
    missing = [c for c in needed_columns if col_map.get(c) is None]
    if missing:
        wb.close()
        raise AutomationError(f"'{os.path.basename(path)}' is missing expected column(s): {missing}")
    return wb, ws, col_map


def _read_lst_cell(path):
    """Returns list of row dicts (one per real cell, base-station-only rows skipped)."""
    needed = list(LST_CELL_DIRECT_COLUMNS.values())
    wb, ws, col_map = _read_source_sheet(path, needed)
    idx = {name: col_map[name] - 1 for name in needed}
    cell_name_idx = idx["Cell Name"]
    max_col = max(idx.values()) + 1

    rows = []
    for row_tuple in ws.iter_rows(
        min_row=SOURCE_DATA_START_ROW, max_row=SOURCE_DATA_START_ROW + 500_000,
        max_col=max_col, values_only=True,
    ):
        if all(v is None for v in row_tuple):
            break
        cell_name_val = row_tuple[cell_name_idx] if cell_name_idx < len(row_tuple) else None
        if cell_name_val is None or str(cell_name_val).strip() == "":
            continue  # base-station-only row, no actual cell here
        row_values = {}
        for template_col, src_col in LST_CELL_DIRECT_COLUMNS.items():
            i = idx[src_col]
            row_values[template_col] = row_tuple[i] if i < len(row_tuple) else None
        rows.append(row_values)
    wb.close()
    return rows


def _build_lookup(path, value_columns):
    """Returns dict: join_key -> {value_column: value, ...}."""
    needed = JOIN_KEY_SOURCE_COLUMNS + value_columns
    wb, ws, col_map = _read_source_sheet(path, needed)
    idx = {name: col_map[name] - 1 for name in needed}
    max_col = max(idx.values()) + 1

    lookup = {}
    for row_tuple in ws.iter_rows(
        min_row=SOURCE_DATA_START_ROW, max_row=SOURCE_DATA_START_ROW + 500_000,
        max_col=max_col, values_only=True,
    ):
        if all(v is None for v in row_tuple):
            break
        key_parts = [row_tuple[idx[c]] if idx[c] < len(row_tuple) else None for c in JOIN_KEY_SOURCE_COLUMNS]
        if any(p is None for p in key_parts):
            continue
        key = "".join(str(p) for p in key_parts)
        lookup[key] = {c: (row_tuple[idx[c]] if idx[c] < len(row_tuple) else None) for c in value_columns}
    wb.close()
    return lookup


def run_pipeline(lst_cell_path, lstpdsch_path, celldlpcpdschpa_path, template_path, output_dir):
    """
    Run the Huawei pipeline and return (output_path, summary_dict). Raises
    AutomationError on any expected, user-fixable problem.
    """
    os.makedirs(output_dir, exist_ok=True)
    summary = {"warnings": []}

    def note(msg):
        summary["warnings"].append(msg)
        log(msg)

    log(f"Reading LST CELL: {os.path.basename(lst_cell_path)}")
    rows_data = _read_lst_cell(lst_cell_path)
    log(f"Read {len(rows_data)} cell rows from LST CELL.")
    summary["source_filename"] = os.path.basename(lst_cell_path)
    summary["rows_read"] = len(rows_data)

    for rv in rows_data:
        rv["__JOIN_KEY__"] = f"{rv.get('Cell ID')}{rv.get('Base Station Name')}"

    log(f"Building lookup from LSTPDSCH: {os.path.basename(lstpdsch_path)}")
    lstpdsch_lookup = _build_lookup(lstpdsch_path, LSTPDSCH_VALUE_COLUMNS)
    log(f"LSTPDSCH lookup has {len(lstpdsch_lookup)} entries.")

    log(f"Building lookup from LST CELLDLPCPDSCHPA: {os.path.basename(celldlpcpdschpa_path)}")
    celldlpcpdschpa_lookup = _build_lookup(celldlpcpdschpa_path, CELLDLPCPDSCHPA_VALUE_COLUMNS)
    log(f"LST CELLDLPCPDSCHPA lookup has {len(celldlpcpdschpa_lookup)} entries.")

    pdsch_misses = 0
    pa_misses = 0
    for rv in rows_data:
        key = rv["__JOIN_KEY__"]
        pdsch_vals = lstpdsch_lookup.get(key)
        if pdsch_vals:
            rv.update(pdsch_vals)
        else:
            for c in LSTPDSCH_VALUE_COLUMNS:
                rv[c] = None
            pdsch_misses += 1
        pa_vals = celldlpcpdschpa_lookup.get(key)
        if pa_vals:
            rv.update(pa_vals)
        else:
            for c in CELLDLPCPDSCHPA_VALUE_COLUMNS:
                rv[c] = None
            pa_misses += 1

    if pdsch_misses:
        note(f"{pdsch_misses} row(s) had no matching Cell ID+Base Station Name in LSTPDSCH -- "
             f"Reference signal power/PB left blank for those.")
    if pa_misses:
        note(f"{pa_misses} row(s) had no matching Cell ID+Base Station Name in LST CELLDLPCPDSCHPA -- "
             f"PA for even power distribution left blank for those.")
    summary["lstpdsch"] = {"entries": len(lstpdsch_lookup), "misses": pdsch_misses}
    summary["celldlpcpdschpa"] = {"entries": len(celldlpcpdschpa_lookup), "misses": pa_misses}

    # Prepare output workbook from template ---------------------------------
    run_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_name = f"{os.path.splitext(summary['source_filename'])[0]}_filled_{run_timestamp}.xlsx"
    output_path = os.path.join(output_dir, output_name)
    shutil.copy(template_path, output_path)

    out_wb = load_workbook(output_path)
    out_ws = out_wb["Sheet1"] if "Sheet1" in out_wb.sheetnames else out_wb.active

    tpl_header_grids = fetch_header_rows(out_ws, TEMPLATE_HEADER_ROWS, max_col=out_ws.max_column)
    tpl_merged = get_merged_ranges(out_ws)
    needed_template_cols = TEMPLATE_DIRECT_COLUMNS + TEMPLATE_LOOKUP_COLUMNS + [
        "Absolute sector", TEMPLATE_JOIN_KEY_COLUMN,
    ]
    tpl_col_map = find_header_columns_fast(
        tpl_header_grids, tpl_merged, TEMPLATE_HEADER_ROWS, needed_template_cols, out_ws.max_column
    )
    missing_in_template = [c for c in needed_template_cols if tpl_col_map.get(c) is None]
    if missing_in_template:
        note("These columns were not found in the template headers and will be skipped: "
             + ", ".join(missing_in_template))

    join_key_col = tpl_col_map.get(TEMPLATE_JOIN_KEY_COLUMN)
    bsn_col = tpl_col_map.get("Base Station Name")
    cell_id_col = tpl_col_map.get("Cell ID")

    abs_sector_col = tpl_col_map.get("Absolute sector")
    cell_name_col = tpl_col_map.get("Cell Name")
    data_columns = TEMPLATE_DIRECT_COLUMNS + TEMPLATE_LOOKUP_COLUMNS

    log(f"Writing into template sheet: '{out_ws.title}' (starting at row {TEMPLATE_DATA_START_ROW})")

    for i, rv in enumerate(rows_data):
        out_row = TEMPLATE_DATA_START_ROW + i
        for template_col in data_columns:
            col = tpl_col_map.get(template_col)
            if col is None:
                continue
            out_ws.cell(row=out_row, column=col, value=rv.get(template_col))

        if join_key_col and cell_id_col and bsn_col:
            cell_id_ref = f"{get_column_letter(cell_id_col)}{out_row}"
            bsn_ref = f"{get_column_letter(bsn_col)}{out_row}"
            out_ws.cell(row=out_row, column=join_key_col, value=f"=CONCATENATE({cell_id_ref},{bsn_ref})")

        if abs_sector_col and cell_name_col:
            cell_name_ref = f"{get_column_letter(cell_name_col)}{out_row}"
            out_ws.cell(row=out_row, column=abs_sector_col, value=absolute_sector_formula(cell_name_ref))

    out_wb.calculation.fullCalcOnLoad = True
    out_wb.save(output_path)
    log(f"\nDone. Output saved to: {output_path}")
    summary["output_path"] = output_path
    summary["output_name"] = output_name

    return output_path, summary
